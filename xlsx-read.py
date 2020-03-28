import openpyxl


plan = openpyxl.load_workbook('D:/Box Sync/Documents/PyCharmProjects/openpyxl/example.xlsx', read_only=True)
plan.get_sheet_by_name('Plan1')
aba = plan.active

print(aba["A1"].value)
print(aba.cell(row=1, column=1).value)

for y in range(1, aba.max_row+1):
    print(y, aba.cell(row=y, column=1).value, aba.cell(row=y, column=2).value, aba.cell(row=y, column=3).value)

print("A última linha da aba é", aba.max_row, " e a última coluna é", aba.max_column)